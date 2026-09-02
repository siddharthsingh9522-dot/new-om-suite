# ==========================================
# OM Automation V2
# api/gst_captcha.py
#
# Captcha-driven GST verification against the real portal
# (services.gst.gov.in), via gst_search.py. This is the fix for GST
# lookups: api/gst_api.py's direct API call can't pass the portal's
# captcha check, so it never returned real data. This module runs an
# actual browser session per user, one GSTIN at a time: show captcha
# -> user types it -> submit -> save result -> show next captcha.
# Session state is kept in-memory, keyed by a Flask session cookie
# ("gst_sid"), same pattern as the standalone tool this was ported
# from.
# ==========================================

import os
import re
import threading
import uuid

from flask import session
from openpyxl import Workbook, load_workbook

from modules.om_automation.config import CAPTCHA_DIR, OUTPUT_DIR, GST_OUTPUT_COLUMNS, GSTIN_HEADERS

from modules.om_automation.excel.generic_reader import GenericReader

from modules.om_automation.gst_search import (
    open_browser, go_to_search_page, fill_gst,
    screenshot_captcha, fill_captcha, click_search,
    get_page_text, looks_like_wrong_captcha, extract_result,
    close_browser,
)

SESSIONS = {}
LOCK = threading.Lock()

_GSTIN_RE = re.compile(r"^[0-9A-Za-z]{15}$")


def is_valid_gstin(value):
    """Structural check only (15 alphanumeric characters) - not a full
    checksum validation."""
    return bool(_GSTIN_RE.match((value or "").strip()))


def _new_sid():
    sid = uuid.uuid4().hex
    session["gst_sid"] = sid
    return sid


def get_state():
    sid = session.get("gst_sid")
    if not sid:
        return None
    return SESSIONS.get(sid)


def _output_path(sid):
    return os.path.join(OUTPUT_DIR, f"GST_Report_{sid}.xlsx")


def _captcha_path(sid):
    return os.path.join(CAPTCHA_DIR, f"{sid}.png")


def _flatten(value):
    """Excel/Calc auto-expands a row's height to fit every embedded
    line break inside a cell - the full-page Raw Text column is full
    of them, which is what made downloaded reports look broken/huge.
    Collapse them to a visible separator instead so every row stays
    one line tall."""
    if isinstance(value, str):
        return value.replace("\r\n", " | ").replace("\n", " | ").replace("\r", " | ")
    return value


def append_result(output_path, result):
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(GST_OUTPUT_COLUMNS)

    ws.append([_flatten(result.get(col, "")) for col in GST_OUTPUT_COLUMNS])
    wb.save(output_path)


def _advance(state, sid):
    """Move to the next GSTIN: fill it in and get a fresh captcha ready."""
    state["current_index"] += 1
    idx = state["current_index"]
    total = len(state["gst_list"])

    if idx >= total:
        close_browser(state["driver"])
        state["driver"] = None
        return {
            "status": "done",
            "total": total,
            "processed": total,
            "output_filename": os.path.basename(state["output_path"]),
        }

    gst = state["gst_list"][idx]
    driver = state["driver"]

    go_to_search_page(driver)
    fill_gst(driver, gst)
    screenshot_captcha(driver, _captcha_path(sid))

    return {
        "status": "waiting_captcha",
        "gst": gst,
        "index": idx,
        "total": total,
    }


def start_session(gst_list):
    """gst_list: list of already-cleaned GSTIN strings."""
    sid = _new_sid()
    driver = open_browser()

    state = {
        "gst_list": gst_list,
        "current_index": -1,
        "driver": driver,
        "output_path": _output_path(sid),
    }
    SESSIONS[sid] = state

    return _advance(state, sid)


def start_single(gstin):
    return start_session([gstin])


def start_excel(file_storage):
    records = GenericReader(file_storage, GSTIN_HEADERS, label="GSTIN").read()

    gst_list = []
    seen = set()
    for r in records:
        value = r["value"].strip().upper()
        if value in seen:
            continue
        seen.add(value)
        gst_list.append(value)

    if not gst_list:
        return {"error": "No GSTIN column found in that file."}

    return start_session(gst_list)


def submit_captcha(captcha_text):
    state = get_state()
    if not state or not state.get("driver"):
        return {"error": "Session expired. Please start again."}, 400

    captcha_text = (captcha_text or "").strip()
    if not captcha_text:
        return {"error": "Please type the captcha text."}, 400

    sid = session["gst_sid"]

    with LOCK:
        driver = state["driver"]
        gst = state["gst_list"][state["current_index"]]

        try:
            fill_captcha(driver, captcha_text)
            click_search(driver)
        except Exception as e:
            return {"error": f"Could not submit captcha: {e}"}, 500

        page_text = get_page_text(driver)

        if looks_like_wrong_captcha(page_text):
            # Same GSTIN, fresh captcha - don't advance the index.
            go_to_search_page(driver)
            fill_gst(driver, gst)
            screenshot_captcha(driver, _captcha_path(sid))
            return {
                "status": "retry_captcha",
                "message": "Captcha didn't match - a new one is ready, try again.",
                "gst": gst,
                "index": state["current_index"],
                "total": len(state["gst_list"]),
            }, 200

        result = extract_result(driver, gst)
        append_result(state["output_path"], result)

        payload = _advance(state, sid)
        payload["last_result"] = {"gst": gst, "remarks": result.get("Remarks", "")}
        if payload["status"] != "done":
            payload["processed"] = state["current_index"]
        return payload, 200


def stop_session():
    state = get_state()
    if not state:
        return {"error": "Session expired."}, 400

    with LOCK:
        if state.get("driver"):
            close_browser(state["driver"])
            state["driver"] = None

        processed = max(state["current_index"], 0)
        total = len(state["gst_list"])
        output_path = state["output_path"]

        return {
            "ok": True,
            "processed": processed,
            "total": total,
            "output_filename": os.path.basename(output_path) if os.path.exists(output_path) else None,
        }, 200


def captcha_image_path():
    sid = session.get("gst_sid")
    if not sid:
        return None
    path = _captcha_path(sid)
    return path if os.path.exists(path) else None
