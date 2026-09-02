# ==========================================
# OM Automation V2
# excel/generic_reader.py
# ==========================================

from openpyxl import load_workbook

from modules.om_automation.core.mapper import get_columns


class GenericReader:
    """
    Like excel.reader.ExcelReader, but matches against
    a caller-supplied header alias list instead of the
    hardcoded CN_HEADERS. Used by the GST Query and Party
    Code Query pages to read GSTIN / Party Code columns.
    """

    def __init__(self, file, header_list, label="Value"):
        self.file = file
        self.header_list = header_list
        self.label = label
        self.records = []
        self.duplicates = set()
        self.seen = set()

    def read(self):

        wb = load_workbook(self.file, data_only=True)

        for sheet in wb.worksheets:

            columns = get_columns(sheet, self.header_list)

            if len(columns) == 0:
                print(f"[WARNING] Sheet '{sheet.title}' has no {self.label} column.")
                continue

            max_row = sheet.max_row

            for row_no in range(2, max_row + 1):

                for col in columns:

                    cell = sheet.cell(row=row_no, column=col["index"] + 1)
                    value = cell.value

                    if value is None:
                        continue

                    text = str(value).strip()

                    if text == "":
                        continue

                    duplicate = False

                    if text in self.seen:
                        duplicate = True
                        self.duplicates.add(text)
                    else:
                        self.seen.add(text)

                    self.records.append({
                        "sheet": sheet.title,
                        "row": row_no,
                        "column": col["header"],
                        "value": text,
                        "duplicate": duplicate
                    })

        return self.records

    def get_duplicates(self):
        return list(self.duplicates)

    def total_records(self):
        return len(self.records)
