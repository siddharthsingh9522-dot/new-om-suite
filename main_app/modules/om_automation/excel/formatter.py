# ==========================================
# OM Automation V2
# excel/formatter.py
# ==========================================

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment


class ExcelFormatter:
    """
    Owns all visual styling for report worksheets.
    excel/writer.py is responsible for structure
    (headers, rows, columns) and calls into this
    class for how things should look.
    """

    def __init__(self):

        self.header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)

        self.success_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        self.error_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        self.alt_fill = PatternFill(fill_type="solid", fgColor="F7F7F7")

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        self.center = Alignment(horizontal="center", vertical="center")

    # ------------------------------------

    def format_header(self, ws):

        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center

    # ------------------------------------

    def format_data_borders(self, ws):
        """Apply a thin border to every data cell (rows below the header)."""

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border

    # ------------------------------------

    def apply_fill(self, ws, fill):
        """Apply a single fill color to every data row (e.g. success/error sheets)."""

        if fill is None:
            return

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill

    # ------------------------------------

    def alternate_rows(self, ws, start_row=2):
        """
        Zebra-stripe data rows using self.alt_fill on every
        other row, for sheets that don't already have a
        success/error fill applied.
        """

        for i, row in enumerate(ws.iter_rows(min_row=start_row)):

            if i % 2 == 1:
                for cell in row:
                    cell.fill = self.alt_fill

    # ------------------------------------

    def auto_width(self, ws):

        from openpyxl.utils import get_column_letter

        for column in ws.columns:

            width = 0
            letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    width = max(width, len(str(cell.value)))

            ws.column_dimensions[letter].width = min(width + 5, 60)


formatter = ExcelFormatter()
