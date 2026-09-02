# ==========================================
# OM Automation V2
# excel/reader.py
# ==========================================

from openpyxl import load_workbook

from modules.om_automation.core.mapper import get_cn_columns


class ExcelReader:

    def __init__(self, file):
        self.file = file
        self.records = []
        self.duplicates = set()
        self.cn_seen = set()

    def read(self):

        wb = load_workbook(self.file, data_only=True)

        for sheet in wb.worksheets:

            cn_columns = get_cn_columns(sheet)

            if len(cn_columns) == 0:
                print(f"[WARNING] Sheet '{sheet.title}' has no CN column.")
                continue

            max_row = sheet.max_row

            for row_no in range(2, max_row + 1):

                for col in cn_columns:

                    cell = sheet.cell(row=row_no, column=col["index"] + 1)
                    value = cell.value

                    if value is None:
                        continue

                    cn = str(value).strip()

                    if cn == "":
                        continue

                    duplicate = False

                    if cn in self.cn_seen:
                        duplicate = True
                        self.duplicates.add(cn)
                    else:
                        self.cn_seen.add(cn)

                    self.records.append({
                        "sheet": sheet.title,
                        "row": row_no,
                        "column": col["header"],
                        "cn": cn,
                        "duplicate": duplicate
                    })

        return self.records

    def get_duplicates(self):
        return list(self.duplicates)

    def total_records(self):
        return len(self.records)
