"""
Customer Search — ported from the standalone customer_webapp into a
Blueprint for the unified OM Suite app. Logic is unchanged; only the
url_prefix ("/customer") and login requirement are new.
"""

import os
import re
import time
import uuid
import threading
import requests
from bs4 import BeautifulSoup
from flask import Blueprint, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.utils import get_column_letter

from auth import login_required

USER_ID = "39942"
BASE_URL = f"https://scmomsanchar.omlogistics.co.in/oracle/query/customer.php?u={USER_ID}"
DELAY_SECONDS = 0.4
TIMEOUT_SECONDS = 30

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Content-Type": "application/x-www-form-urlencoded",
}

RESULT_COLUMNS = [
    "CCode", "CNAME", "Address", "Booking Location", "Acc Loc", "Bill Loc",
    "Verify", "Type", "GST NO", "Status", "Master Code", "Kom", "Document",
    "First Booking date", "Vendor Code", "CUSTOMER MOBILE NO", "MERGE CODE",
    "DELIVERY CODE", "RETAIL TRANSFER DATE", "CONT RATE",
    "CUSTOMER ENTER BY", "CUSTOMER ENTER DATE",
]

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def clean_cell_value(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "customer_uploads")
RESULTS_DIR = os.path.join(BASE_DIR, "data", "customer_results")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)

customer_bp = Blueprint(
    "customer", __name__,
    url_prefix="/customer",
    template_folder="../templates/customer",
)


def search_customer(session, cname="", ccode="", lcode="", blcode="", mcode="", custgst=""):
    payload = {
        "cname": cname, "ccode": ccode, "lcode": lcode,
        "blcode": blcode, "mcode": mcode, "custgst": custgst,
        "submit": "Search",
    }
    resp = session.post(BASE_URL, data=payload, headers=HEADERS, timeout=TIMEOUT_SECONDS)
    resp.raise_for_status()

    no_data_found = "no data found" in resp.text.lower()

    soup = BeautifulSoup(resp.text, "html.parser")
    table = soup.find("table")
    if table is None:
        return [], no_data_found

    rows = []
    tbody = table.find("tbody") or table
    for tr in tbody.find_all("tr"):
        cells = tr.find_all("td")
        if not cells:
            continue
        values = [c.get_text(strip=True) for c in cells]
        if len(values) < len(RESULT_COLUMNS):
            values += [""] * (len(RESULT_COLUMNS) - len(values))
        elif len(values) > len(RESULT_COLUMNS):
            values = values[: len(RESULT_COLUMNS)]
        rows.append(dict(zip(RESULT_COLUMNS, values)))
    return rows, no_data_found


def search_with_retry(session, code, job=None):
    try:
        rows, no_data_found = search_customer(session, blcode=code)
    except requests.RequestException:
        rows, no_data_found = [], False

    if rows or no_data_found:
        return rows

    for _ in range(15):
        if job is not None and job.get("stop_requested"):
            return rows
        time.sleep(1)

    try:
        rows, _ = search_customer(session, blcode=code)
    except requests.RequestException:
        rows = []
    return rows


def find_code_column(ws):
    header_row = [c.value for c in ws[1]]
    keywords = ["ccode", "customer code", "cust code", "code", "customer_code"]
    for idx, header in enumerate(header_row, start=1):
        if header and str(header).strip().lower() in keywords:
            return idx, True

    best_col, best_count = None, 0
    for col in range(1, ws.max_column + 1):
        count = 0
        for row in range(1, min(ws.max_row, 50) + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip().replace(".", "", 1).isdigit():
                count += 1
        if count > best_count:
            best_count, best_col = count, col

    if best_col is None:
        raise ValueError("Could not detect a column containing customer codes.")

    first_val = ws.cell(row=1, column=best_col).value
    has_header = not (first_val is not None and str(first_val).strip().replace(".", "", 1).isdigit())
    return best_col, has_header


def read_customer_codes(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    col_idx, has_header = find_code_column(ws)
    start_row = 2 if has_header else 1

    codes = []
    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val is None or str(val).strip() == "":
            continue
        s = str(val).strip()
        if s.endswith(".0"):
            s = s[:-2]
        codes.append(s)
    return codes


def write_results_excel(path, all_results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.append(["Searched Code"] + RESULT_COLUMNS)
    for col_idx in range(1, len(RESULT_COLUMNS) + 2):
        ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)

    not_found = []
    for searched_code, rows in all_results:
        searched_code = clean_cell_value(searched_code)
        if not rows:
            ws.append([searched_code] + ["NOT FOUND"] + [""] * (len(RESULT_COLUMNS) - 1))
            not_found.append(searched_code)
            continue
        for row in rows:
            ws.append([searched_code] + [clean_cell_value(row[c]) for c in RESULT_COLUMNS])

    for col_idx in range(1, len(RESULT_COLUMNS) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    if not_found:
        ws2 = wb.create_sheet("Not Found")
        ws2.append(["Customer Code"])
        for code in not_found:
            ws2.append([code])

    wb.save(path)


@customer_bp.route("/")
@login_required
def index():
    return render_template("customer/index.html", columns=RESULT_COLUMNS)


@customer_bp.route("/api/manual-search", methods=["POST"])
@login_required
def manual_search():
    data = request.get_json(force=True)
    session = requests.Session()
    try:
        rows, _ = search_customer(
            session,
            cname=data.get("cname", ""), ccode=data.get("ccode", ""),
            lcode=data.get("lcode", ""), blcode=data.get("blcode", ""),
            mcode=data.get("mcode", ""), custgst=data.get("custgst", ""),
        )
        return jsonify({"ok": True, "rows": rows, "columns": RESULT_COLUMNS})
    except requests.RequestException as e:
        return jsonify({"ok": False, "error": str(e)}), 502


JOBS = {}
JOBS_LOCK = threading.Lock()


def _write_job_output(job_id, all_results):
    output_filename = f"customer_search_results_{job_id}.xlsx"
    output_path = os.path.join(RESULTS_DIR, output_filename)
    write_results_excel(output_path, all_results)
    return f"/customer/download/{output_filename}"


def run_bulk_job(job_id, input_path):
    job = JOBS[job_id]
    try:
        codes = read_customer_codes(input_path)
    except Exception as e:
        job["status"] = "error"
        job["error"] = f"Could not read Excel: {e}"
        return

    if not codes:
        job["status"] = "error"
        job["error"] = "No customer codes found in the file"
        return

    job["total"] = len(codes)
    job["status"] = "running"

    session = requests.Session()
    all_results = []
    stopped = False

    try:
        for code in codes:
            if job["stop_requested"]:
                stopped = True
                break
            rows = search_with_retry(session, code, job=job)
            all_results.append((code, rows))
            job["done"] += 1
            if rows:
                job["found"] += 1
            else:
                job["not_found"] += 1
            if job["stop_requested"]:
                stopped = True
                break
            time.sleep(DELAY_SECONDS)

        job["download_url"] = _write_job_output(job_id, all_results)
        job["status"] = "stopped" if stopped else "done"
    except Exception as e:
        job["status"] = "error"
        job["error"] = f"Bulk search failed: {e}"


@customer_bp.route("/api/bulk-search/start", methods=["POST"])
@login_required
def bulk_search_start():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "No file selected"}), 400

    job_id = uuid.uuid4().hex[:10]
    input_path = os.path.join(UPLOAD_DIR, f"{job_id}_{file.filename}")
    file.save(input_path)

    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "starting", "total": 0, "done": 0, "found": 0,
            "not_found": 0, "stop_requested": False, "error": None,
            "download_url": None,
        }

    thread = threading.Thread(target=run_bulk_job, args=(job_id, input_path), daemon=True)
    thread.start()

    return jsonify({"ok": True, "job_id": job_id})


@customer_bp.route("/api/bulk-search/status/<job_id>")
@login_required
def bulk_search_status(job_id):
    job = JOBS.get(job_id)
    if job is None:
        return jsonify({"ok": False, "error": "Unknown job id"}), 404
    return jsonify({"ok": True, **job})


@customer_bp.route("/api/bulk-search/stop/<job_id>", methods=["POST"])
@login_required
def bulk_search_stop(job_id):
    job = JOBS.get(job_id)
    if job is None:
        return jsonify({"ok": False, "error": "Unknown job id"}), 404
    job["stop_requested"] = True
    return jsonify({"ok": True})


@customer_bp.route("/download/<filename>")
@login_required
def download_result(filename):
    path = os.path.join(RESULTS_DIR, filename)
    return send_file(path, as_attachment=True, download_name="customer_search_results.xlsx")
