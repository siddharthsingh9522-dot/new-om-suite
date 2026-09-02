import os
import io
import csv
import json
import hashlib
import imaplib
import difflib
from email.utils import getaddresses
import re
import smtplib
import ssl
from email.message import EmailMessage
from datetime import datetime

import requests
import pandas as pd
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session, abort, g
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

from directory_parser import (
    parse_directory, parse_report, build_branches_with_email,
    pick_branch_email, pick_branch_role_contacts,
)
import bill_parser
import previous_mail
import db
from auth import login_required, admin_required, current_user, get_csrf_token, validate_csrf

from modules.branch import branch_bp
from modules.customer import customer_bp
from modules.om_automation_bp import om_bp

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "change-me")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    MAX_CONTENT_LENGTH=20 * 1024 * 1024,  # 20MB upload cap
)

ALLOWED_UPLOAD_EXTENSIONS = {".xls", ".xlsx"}


def _validate_upload_extension(file_storage):
    ext = os.path.splitext(file_storage.filename or "")[1].lower()
    return ext in ALLOWED_UPLOAD_EXTENSIONS

db.init_db()

app.register_blueprint(branch_bp)
app.register_blueprint(customer_bp)
app.register_blueprint(om_bp)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
TBB_LOG_FILE = os.path.join(DATA_DIR, "tbb_send_log.csv")
BILL_LOG_FILE = os.path.join(DATA_DIR, "bill_send_log.csv")
PREVIOUS_LOG_FILE = os.path.join(DATA_DIR, "previous_send_log.csv")
DIRECTORY_CACHE = os.path.join(DATA_DIR, "directory_cache.xlsx")

os.makedirs(UPLOAD_DIR, exist_ok=True)

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# ---------- SMTP / IMAP / AI config ----------
# Ye ab admin/.env se globally NAHI aate — har user apni khud ki Mail (SMTP+IMAP)
# aur AI (Gemini/Anthropic) details "/settings" page se save karta hai, DB mein
# encrypted rehti hain (crypto_utils), aur har request ke shuru mein
# load_request_credentials() (neeche) unhe g.mail_cfg / g.ai_cfg mein daal deta hai.
# Model names abhi bhi ek shared default rakhe hain (ye secret nahi hain, sirf
# konsa model use karna hai — user chahe to future mein per-user bhi ban sakte hain).
ANTHROPIC_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-5").strip()
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.1-flash-lite").strip()

# GR Auto Modification System apna alag Render service hai, apna alag login hai —
# ye sirf uska URL hai taaki dashboard par ek tile/link dikha sakein. Admin ise
# GR Auto ke deploy hone ke baad uska Render URL daal kar set karta hai.
GR_AUTO_URL = os.environ.get("GR_AUTO_URL", "").strip()

TBB_EMAIL_TEMPLATE_INTRO = (
    "Dear sir,\n\n"
    "Please check the GR. If this GR belongs to a TBB customer, kindly provide the "
    "correct TBB Party Code and modify the same accordingly.\n\n"
    "Please confirm once updated.\n\n"
)

# Har user ki organization directory (branch->email) shared reference data hai — safe hai
# share karna kyunki ye kisi ki personal business data nahi, sirf company ki directory hai.
# TBB/Bill workflow data ab db.py ke workflow_data table mein user_id ke saath persist hoti hai
# (in-memory nahi) — isliye app restart hone par bhi user ka data uske account se juda rehta hai.
STATE = {
    "directory": {},
}


@app.before_request
def load_request_credentials():
    """Har request ke shuru mein logged-in user ki apni Mail/AI settings DB se
    load karke g.mail_cfg / g.ai_cfg mein daal deta hai. Koi bhi function ab
    se admin ka shared SMTP/API key nahi, balki ye per-request values use karta
    hai. Logged-out request ya user ne kabhi /settings bhara hi nahi ho, to
    dono khaali/'not configured' state mein rehte hain."""
    user = current_user()
    if user:
        creds = db.get_user_credentials(user["id"])
        g.mail_cfg = creds["mail"]
        g.ai_cfg = creds["ai"]
    else:
        g.mail_cfg = {"configured": False}
        g.ai_cfg = {"provider": "none"}


def get_directory():
    """Directory cache file se load karta hai agar in-memory khali ho (jaise restart ke baad)."""
    if STATE["directory"]:
        return STATE["directory"]
    if os.path.exists(DIRECTORY_CACHE):
        STATE["directory"] = parse_directory(DIRECTORY_CACHE)
    return STATE["directory"]


def user_upload_dir(user_id):
    d = os.path.join(UPLOAD_DIR, str(int(user_id)))
    os.makedirs(d, exist_ok=True)
    return d


def compute_fingerprint(module, branch_code, email, content_text):
    """Deterministic hash of branch+recipient+message content. Same content -> same
    fingerprint -> sync_operations() treats it as the same operation (idempotent).
    Different content (e.g. a new day's data) -> different fingerprint -> fresh operation."""
    raw = f"{module}:{branch_code}:{email}:{content_text}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ---------------- Shared helpers ----------------

def log_send(log_file, branch_code, branch_name, email, status, detail="", username=""):
    is_new = not os.path.exists(log_file)
    with open(log_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(["timestamp", "username", "branch_code", "branch_name", "email", "status", "detail"])
        writer.writerow([datetime.now().isoformat(timespec="seconds"), username, branch_code, branch_name, email, status, detail])


def _smtp_error_message(exc):
    """SMTP errors ko user-friendly message mein convert karta hai."""
    text = str(exc)
    cfg = g.mail_cfg
    if "553" in text and "not owned" in text.lower():
        return (
            "SMTP sender rejected: aapke SMTP username ke paas 'From' email use karne ki "
            f"permission nahi hai. Configured sender={cfg.get('smtp_from_email')!r}, "
            f"login={cfg.get('smtp_user')!r}. /settings mein jaake apni Mail Settings check karein."
        )
    if "535" in text or "authentication" in text.lower():
        return "SMTP authentication failed: /settings mein apna SMTP username/password check karein."
    return text


def parse_recipients(value):
    """Accept comma, semicolon or newline separated addresses."""
    if not value:
        return []
    raw = re.split(r"[,;\n]+", str(value))
    out, seen = [], set()
    for item in raw:
        email = item.strip().lower()
        if not email:
            continue
        if not EMAIL_RE.match(email):
            raise ValueError(f"Invalid email address: {email}")
        if email not in seen:
            seen.add(email); out.append(email)
    if not out:
        raise ValueError("At least one valid recipient email is required.")
    return out


def _append_to_sent(msg):
    """Best-effort copy to IMAP Sent folder, current user ki apni IMAP config se.
    SMTP success is never reversed if this fails."""
    cfg = g.mail_cfg
    if not cfg.get("imap_host") or not cfg.get("imap_user") or not cfg.get("imap_pass"):
        return {"saved": False, "reason": "IMAP not configured"}
    try:
        with imaplib.IMAP4_SSL(cfg["imap_host"], int(cfg.get("imap_port") or 993)) as imap:
            imap.login(cfg["imap_user"], cfg["imap_pass"])
            typ, _ = imap.append(cfg.get("imap_sent_folder") or "Sent", "\\Seen",
                                  imaplib.Time2Internaldate(__import__('time').time()), msg.as_bytes())
            if typ != "OK":
                return {"saved": False, "reason": "IMAP append rejected"}
        return {"saved": True}
    except Exception as exc:
        return {"saved": False, "reason": str(exc)}


def _send_smtp(msg, envelope_recipients):
    cfg = g.mail_cfg
    if not cfg.get("smtp_host") or not cfg.get("smtp_user") or not cfg.get("smtp_pass"):
        raise RuntimeError("Aapki Mail Settings abhi configure nahi hain. Pehle /settings mein jaake "
                            "apna SMTP Host, Username aur Password save karein, phir dobara try karein.")
    recipients = list(dict.fromkeys(envelope_recipients))
    context = ssl.create_default_context()
    smtp_port = int(cfg.get("smtp_port") or 587)
    from_email = cfg.get("smtp_from_email") or cfg["smtp_user"]
    try:
        if smtp_port == 465:
            with smtplib.SMTP_SSL(cfg["smtp_host"], smtp_port, context=context, timeout=30) as server:
                server.login(cfg["smtp_user"], cfg["smtp_pass"])
                refused = server.send_message(msg, from_addr=from_email, to_addrs=recipients)
        else:
            with smtplib.SMTP(cfg["smtp_host"], smtp_port, timeout=30) as server:
                server.ehlo(); server.starttls(context=context); server.ehlo(); server.login(cfg["smtp_user"], cfg["smtp_pass"])
                refused = server.send_message(msg, from_addr=from_email, to_addrs=recipients)
        if refused:
            raise RuntimeError(f"SMTP recipient refused: {refused}")
        return _append_to_sent(msg)
    except Exception as exc:
        raise RuntimeError(_smtp_error_message(exc)) from exc


def _call_anthropic(prompt):
    api_key = g.ai_cfg.get("anthropic_api_key", "")
    if not api_key:
        raise RuntimeError("Aapki Anthropic API key /settings mein configured nahi hai.")
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": ANTHROPIC_MODEL,
            "max_tokens": 300,
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=30,
    )
    if not resp.ok:
        raise RuntimeError(f"Anthropic HTTP {resp.status_code}: {resp.text[:800]}")
    data = resp.json()
    text = "".join(
        block.get("text", "")
        for block in data.get("content", [])
        if isinstance(block, dict)
    ).strip()
    if not text:
        raise RuntimeError(f"Anthropic returned empty response: {data}")
    return text


def _call_gemini(prompt):
    """gemini-3.1-flash-lite via current Interactions REST API."""
    api_key = g.ai_cfg.get("gemini_api_key", "")
    if not api_key:
        raise RuntimeError("Aapki Gemini API key /settings mein configured nahi hai.")

    url = "https://generativelanguage.googleapis.com/v1beta/interactions"
    schema = {
        "type": "object",
        "properties": {
            "status": {"type": "string", "enum": ["ok", "review"]},
            "reason": {"type": "string"},
        },
        "required": ["status", "reason"],
    }
    payload = {
        "model": GEMINI_MODEL,
        "input": prompt,
        "response_format": {
            "type": "text",
            "mime_type": "application/json",
            "schema": schema,
        },
    }
    resp = requests.post(
        url,
        headers={
            "x-goog-api-key": api_key,
            "content-type": "application/json",
        },
        json=payload,
        timeout=30,
    )
    if not resp.ok:
        raise RuntimeError(f"Gemini HTTP {resp.status_code}: {resp.text[:1000]}")

    data = resp.json()
    # Current Interactions API exposes model output as steps[].content[].text.
    text_parts = []
    for step in data.get("steps", []):
        if step.get("type") == "model_output":
            for item in step.get("content", []):
                if isinstance(item, dict) and item.get("type") == "text":
                    text_parts.append(item.get("text", ""))
    text = "".join(text_parts).strip()

    if not text:
        # Compatibility fallback for any response exposing output_text.
        text = str(data.get("output_text", "")).strip()
    if not text:
        raise RuntimeError(f"Gemini returned empty output: {data}")
    return text


def _parse_ai_json(text):
    text = (text or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*```$", "", text).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end < start:
        raise ValueError(f"AI ne valid JSON return nahi kiya: {text[:500]}")
    return json.loads(text[start:end + 1])


def run_ai_check(prompt):
    """Current user ke configured provider se AI check. Key missing ho to
    explicit unavailable status deta hai (/settings mein bharne ka pointer)."""
    provider = g.ai_cfg.get("provider", "none")
    if provider == "none":
        return {
            "status": "unavailable",
            "reason": "AI API key /settings mein configured nahi hai. Manual verification required.",
        }
    try:
        text = _call_gemini(prompt) if provider == "gemini" else _call_anthropic(prompt)
        parsed = _parse_ai_json(text)
        status = str(parsed.get("status", "review")).lower()
        if status not in ("ok", "review"):
            status = "review"
        return {"status": status, "reason": str(parsed.get("reason", "")).strip()}
    except Exception as exc:
        return {
            "status": "ai_error",
            "reason": f"AI check failed ({provider}): {exc}",
        }


def build_admin_check_prompt(module_label, branch_code, branch_name, region, email, contact, designation, item_count, sample_items):
    return f"""Tum ek logistics company (Om Logistics) ke internal admin ho jo "{module_label}" mail ko
bhejne se pehle check karta hai.

Branch Code: {branch_code}
Branch Naam: {branch_name}
Region: {region}
Bheja jaane wala email: {email}
Contact: {contact or 'N/A'} — Designation: {designation or 'N/A'}
Item Count: {item_count}
Sample: {json.dumps(sample_items, ensure_ascii=False, default=str)}

Check karo: (1) email format/domain sahi hai, (2) designation is kaam ke liye plausible hai
(Security Guard/Cook/Driver jaisi role galat lagegi), (3) data garbage to nahi.

Sirf JSON return karo, kuch aur text nahi:
{{"status": "ok" ya "review", "reason": "chhoti Hindi/Hinglish line"}}"""


# ================== TBB PARTY MAIL — helpers ==================

def tbb_build_subject(branch):
    """Naya format: 'PROVIDE THE TBB CODE & MODIFY THE SAME............. GR:-<pehla>/<aakhri> M/S- <PARTY NAME>'"""
    grs = branch.get("grs", [])
    cns = [str(g.get("cn", "")).strip() for g in grs if str(g.get("cn", "")).strip()]
    if len(cns) >= 2:
        gr_part = f"{cns[0]}/{cns[-1]}"
    else:
        gr_part = cns[0] if cns else ""

    party_names = []
    for g in grs:
        pn = str(g.get("party_name", "")).strip()
        if pn and pn.lower() != "nan" and pn not in party_names:
            party_names.append(pn)
    party_part = " & ".join(party_names)

    return f"PROVIDE THE TBB CODE & MODIFY THE SAME............. GR:-{gr_part} M/S- {party_part}"


def tbb_build_body(branch):
    """Plain-text fallback (jo HTML na dekh paye unke liye) — tab-separated table."""
    body = TBB_EMAIL_TEMPLATE_INTRO
    body += "B.Code \tBRANCH \tCN_NO \tPARTY CODE \tPARTY NAME\n"
    for g in branch["grs"]:
        body += (
            f"{branch['branch_code']} \t{branch.get('branch_name_sheet','')} \t"
            f"{g['cn']} \t{g['party_code']} \t*{g['party_name']}*\n"
        )
    body += "\nRegards,\n" + (g.mail_cfg.get("smtp_from_name") or "")
    return body


def _html_escape(val):
    return (str(val or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def tbb_build_body_html(branch):
    """Proper bordered HTML table — har column apni jagah (B.Code/BRANCH/CN_NO/PARTY CODE/PARTY NAME),
    PARTY NAME bold + red."""
    intro_html = "".join(f"<p>{_html_escape(line)}</p>" for line in TBB_EMAIL_TEMPLATE_INTRO.strip("\n").split("\n\n"))

    headers = ["B.Code", "BRANCH", "CN_NO", "PARTY CODE", "PARTY NAME"]
    th_style = ("background:#1F4E78;color:#fff;padding:6px 10px;font-size:12px;"
                "border:1px solid #333;text-align:left;")
    td_style = "padding:6px 10px;font-size:12px;border:1px solid #333;"

    header_row = "".join(f'<th style="{th_style}">{_html_escape(h)}</th>' for h in headers)

    body_rows = []
    for i, g in enumerate(branch["grs"]):
        bg = "#F2F2F2" if i % 2 == 0 else "#FFFFFF"
        row_td = f'style="{td_style}background:{bg};"'
        party_td = f'style="{td_style}background:{bg};color:#C00000;font-weight:bold;"'
        body_rows.append(
            f"<tr>"
            f'<td {row_td}>{_html_escape(branch["branch_code"])}</td>'
            f'<td {row_td}>{_html_escape(branch.get("branch_name_sheet",""))}</td>'
            f'<td {row_td}>{_html_escape(g["cn"])}</td>'
            f'<td {row_td}>{_html_escape(g["party_code"])}</td>'
            f'<td {party_td}>{_html_escape(g["party_name"])}</td>'
            f"</tr>"
        )

    table = (
        '<table style="border-collapse:collapse;border:1px solid #333;margin-top:6px;">'
        f"<thead><tr>{header_row}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table>"
    )

    footer = f"<p>Regards,<br>{_html_escape(g.mail_cfg.get('smtp_from_name') or '')}</p>"
    return f"<html><body style='font-family:Arial,sans-serif;font-size:13px;'>{intro_html}{table}{footer}</body></html>"


def tbb_build_attachment_xlsx(branch):
    """CSV ki jagah ab proper .xlsx attachment banata hai (openpyxl)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TBB Details"
    headers = ["Branch Code", "Branch Name", "Region", "CN No", "Party Code", "Party Name", "CN Date"]
    ws.append(headers)
    for g in branch["grs"]:
        ws.append([
            branch["branch_code"], branch.get("branch_name_sheet", ""), branch.get("region", ""),
            g["cn"], g["party_code"], g["party_name"], g["cn_date"],
        ])
    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _apply_recipients(msg, to_value, cc_value="", bcc_value=""):
    to_list = parse_recipients(to_value)
    cc_list = parse_recipients(cc_value) if cc_value else []
    bcc_list = parse_recipients(bcc_value) if bcc_value else []
    if cc_list: msg["Cc"] = ", ".join(cc_list)
    # Bcc deliberately not added as a visible header.
    msg["To"] = ", ".join(to_list)
    return to_list, cc_list, bcc_list


def tbb_send_email(branch, to_email, cc="", bcc=""):
    msg = EmailMessage(); msg["Subject"] = tbb_build_subject(branch)
    msg["From"] = f"{g.mail_cfg.get('smtp_from_name') or ''} <{g.mail_cfg.get('smtp_from_email') or ''}>"
    to_list, cc_list, bcc_list = _apply_recipients(msg, to_email, cc, bcc)
    msg.set_content(tbb_build_body(branch))
    msg.add_alternative(tbb_build_body_html(branch), subtype="html")
    xlsx_bytes = tbb_build_attachment_xlsx(branch)
    filename = f"TBB_{branch['branch_code']}_{branch.get('branch_name_sheet','')}".replace(" ", "_") + ".xlsx"
    msg.add_attachment(xlsx_bytes, maintype="application",
                        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
    sent_copy = _send_smtp(msg, to_list + cc_list + bcc_list)
    return {"recipients": to_list, "cc": cc_list, "bcc_count": len(bcc_list), "sent_copy": sent_copy}


def bill_send_email(branch, to_email, cc="", bcc=""):
    msg = EmailMessage(); msg["Subject"] = bill_parser.build_subject(branch)
    msg["From"] = f"{g.mail_cfg.get('smtp_from_name') or ''} <{g.mail_cfg.get('smtp_from_email') or ''}>"
    to_list, cc_list, bcc_list = _apply_recipients(msg, to_email, cc, bcc)
    msg.set_content(bill_parser.build_email_text(branch)); msg.add_alternative(bill_parser.build_email_html(branch), subtype="html")
    sent_copy = _send_smtp(msg, to_list + cc_list + bcc_list)
    return {"recipients": to_list, "cc": cc_list, "bcc_count": len(bcc_list), "sent_copy": sent_copy}


def previous_build_attachment_csv(group):
    buf = io.StringIO()
    writer = csv.writer(buf)
    cols = group.get("columns", [])
    writer.writerow(cols)
    for row in group.get("rows", []):
        writer.writerow([row.get(c, "") for c in cols])
    return buf.getvalue().encode("utf-8")


def previous_send_email(group, to_email, cc="", bcc=""):
    """Previous Mail & Format module ke liye generic sender — subject/body group mein
    pehle se render ho chuke hain (template + is group ke data se)."""
    msg = EmailMessage()
    msg["Subject"] = group.get("subject", "")
    msg["From"] = f"{g.mail_cfg.get('smtp_from_name') or ''} <{g.mail_cfg.get('smtp_from_email') or ''}>"
    to_list, cc_list, bcc_list = _apply_recipients(msg, to_email, cc, bcc)
    msg.set_content(group.get("body", ""))
    if group.get("rows"):
        csv_bytes = previous_build_attachment_csv(group)
        filename = f"PreviousMail_{group.get('key','')}".replace(" ", "_") + ".csv"
        msg.add_attachment(csv_bytes, maintype="text", subtype="csv", filename=filename)
    sent_copy = _send_smtp(msg, to_list + cc_list + bcc_list)
    return {"recipients": to_list, "cc": cc_list, "bcc_count": len(bcc_list), "sent_copy": sent_copy}


def build_previous_template_prompt(old_subject, old_body, sample_row, columns):
    return f"""Tum ek email-template extraction assistant ho. Neeche ek PURANI email ka
subject aur body diya hai, aur NAYE DATA ka ek sample row (jisse naya mail banana hai).

PURANI EMAIL — Subject:
{old_subject}

PURANI EMAIL — Body:
{old_body[:3000]}

NAYE DATA ke columns: {json.dumps(columns, ensure_ascii=False)}
NAYE DATA ka ek sample row: {json.dumps(sample_row, ensure_ascii=False, default=str)}

Kaam: Purani email ke subject aur body ko ek TEMPLATE mein convert karo, jahan jo bhi
specific values (jaise party name, bill number, branch name, date, amount) the unko
{{{{Column Name}}}} placeholders se replace karo — sirf un columns ka naam use karo jo
"NAYE DATA ke columns" list mein di hai. Baaki fixed text (greeting, instructions,
signature) bilkul waisa hi rakho jaisa tha.

Sirf JSON return karo, kuch aur text nahi:
{{"subject_template": "...", "body_template": "...", "confidence": 0-100, "notes": "Hindi/Hinglish mein chhoti si explanation ki kya-kya replace kiya"}}"""


def run_previous_template_ai(old_subject, old_body, sample_row, columns):
    provider = g.ai_cfg.get("provider", "none")
    if provider == "none":
        return {"subject_template": old_subject, "body_template": old_body, "confidence": 0,
                "notes": "AI /settings mein configure nahi hai — purani mail hoobahu copy ki gayi hai, aapko manually {{Column}} placeholders lagane honge."}
    prompt = build_previous_template_prompt(old_subject, old_body, sample_row, columns)
    try:
        text = _call_gemini(prompt) if provider == "gemini" else _call_anthropic(prompt)
        text = text.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(text)
        parsed.setdefault("subject_template", old_subject)
        parsed.setdefault("body_template", old_body)
        parsed.setdefault("confidence", 50)
        parsed.setdefault("notes", "")
        return parsed
    except Exception as e:
        return {"subject_template": old_subject, "body_template": old_body, "confidence": 0,
                "notes": f"AI template extraction fail hui: {e}. Manually placeholders edit kar lein."}


# ---------------- Auth routes ----------------

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return render_template("signup.html", csrf_token=get_csrf_token(), error=None)

    if not validate_csrf(request.form.get("csrf_token")):
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Form expire ho gaya, dobara try karein."), 400

    name = request.form.get("name", "").strip()
    username = request.form.get("username", "").strip().lower()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")
    confirm = request.form.get("confirm_password", "")

    if not (name and username and email and password):
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Sabhi fields bharna zaroori hai."), 400
    if len(username) < 3 or not re.match(r"^[a-z0-9_.]+$", username):
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Username sirf lowercase letters/numbers/./_ ka ho sakta hai (3+ characters)."), 400
    if not EMAIL_RE.match(email):
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Email valid nahi hai."), 400
    if len(password) < 8:
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Password kam se kam 8 characters ka ho."), 400
    if password != confirm:
        return render_template("signup.html", csrf_token=get_csrf_token(), error="Password match nahi kar raha."), 400

    try:
        role, status = db.create_user(name, username, email, password)
    except ValueError as e:
        return render_template("signup.html", csrf_token=get_csrf_token(), error=str(e)), 400

    db.record_audit(username, "USER_SIGNUP", f"role={role} status={status}")

    if status == "active":
        return render_template("signup_done.html", message="Account ban gaya (pehla user hone ki wajah se aap admin ban gaye hain). Ab login karein.")
    return render_template("signup_done.html", message="Account ban gaya hai — admin approval ka wait karein. Approve hone ke baad hi login ho payega.")


def _safe_next_url(candidate):
    """Open-redirect protection: only allow same-site relative paths."""
    if not candidate:
        return url_for("dashboard")
    if candidate.startswith("/") and not candidate.startswith("//") and "\\" not in candidate:
        return candidate
    return url_for("dashboard")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", csrf_token=get_csrf_token(), error=None, next=request.args.get("next", ""))

    if not validate_csrf(request.form.get("csrf_token")):
        return render_template("login.html", csrf_token=get_csrf_token(), error="Form expire ho gaya, dobara try karein.", next=""), 400

    username = request.form.get("username", "").strip().lower()
    password = request.form.get("password", "")
    next_url = _safe_next_url(request.form.get("next"))

    user = db.get_user_by_username(username)
    if not user:
        db.record_audit(username, "LOGIN_FAILURE", "user not found")
        return render_template("login.html", csrf_token=get_csrf_token(), error="Username ya password galat hai.", next=next_url), 401

    if db.is_locked(user):
        db.record_audit(username, "LOGIN_BLOCKED", "account locked")
        return render_template("login.html", csrf_token=get_csrf_token(),
                                error=f"Bahut zyada galat attempts ho gaye — {db.LOCKOUT_MINUTES} minute baad try karein.",
                                next=next_url), 401

    if not db.verify_password(user, password):
        db.register_failed_attempt(username)
        db.record_audit(username, "LOGIN_FAILURE", "wrong password")
        return render_template("login.html", csrf_token=get_csrf_token(), error="Username ya password galat hai.", next=next_url), 401

    if user["status"] == "pending":
        return render_template("login.html", csrf_token=get_csrf_token(),
                                error="Aapka account abhi admin approval ka wait kar raha hai.", next=next_url), 403
    if user["status"] == "disabled":
        db.record_audit(username, "LOGIN_BLOCKED", "account disabled")
        return render_template("login.html", csrf_token=get_csrf_token(), error="Aapka account disable kar diya gaya hai.", next=next_url), 403

    db.reset_failed_attempts(username)
    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    db.record_audit(username, "LOGIN_SUCCESS")
    return redirect(next_url)


@app.route("/logout", methods=["POST"])
def logout():
    user = current_user()
    if user:
        db.record_audit(user["username"], "LOGOUT")
    session.clear()
    return redirect(url_for("login"))


# ---------------- Admin routes ----------------

@app.route("/admin/users", methods=["GET"])
@admin_required
def admin_users():
    return render_template("admin_users.html", users=db.list_users(), csrf_token=get_csrf_token(), me=current_user())


@app.route("/admin/users/<int:user_id>/<action>", methods=["POST"])
@admin_required
def admin_user_action(user_id, action):
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    target = db.get_user_by_id(user_id)
    if not target:
        abort(404)
    me = current_user()
    if action == "approve":
        db.set_user_status(user_id, "active")
    elif action == "disable":
        if target["id"] == me["id"]:
            abort(400)
        db.set_user_status(user_id, "disabled")
    elif action == "enable":
        db.set_user_status(user_id, "active")
    elif action == "make_admin":
        db.set_user_role(user_id, "admin")
    elif action == "make_user":
        if target["id"] == me["id"]:
            abort(400)
        db.set_user_role(user_id, "user")
    else:
        abort(400)
    db.record_audit(me["username"], "ADMIN_ACTION", f"{action} on {target['username']}")
    return redirect(url_for("admin_users"))


# ---------------- My Settings (per-user Mail + AI credentials) ----------------

@app.route("/settings", methods=["GET"])
@login_required
def user_settings():
    return render_template("settings.html", me=current_user(), mail=g.mail_cfg, ai=g.ai_cfg,
                            csrf_token=get_csrf_token(), saved=request.args.get("saved", ""),
                            error=request.args.get("error", ""))


@app.route("/settings/mail", methods=["POST"])
@login_required
def save_mail_settings():
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    user = current_user()
    f = request.form
    try:
        smtp_port = int(f.get("smtp_port") or 587)
        imap_port = int(f.get("imap_port") or 993)
    except ValueError:
        return redirect(url_for("user_settings", error="Port number valid nahi hai."))
    db.save_mail_credentials(
        user["id"],
        smtp_host=(f.get("smtp_host") or "").strip(),
        smtp_port=smtp_port,
        smtp_user=(f.get("smtp_user") or "").strip(),
        smtp_pass=f.get("smtp_pass") or "",  # blank = purana password retain
        smtp_from_email=(f.get("smtp_from_email") or "").strip(),
        smtp_from_name=(f.get("smtp_from_name") or "").strip(),
        imap_host=(f.get("imap_host") or "").strip(),
        imap_port=imap_port,
        imap_user=(f.get("imap_user") or "").strip(),
        imap_pass=f.get("imap_pass") or "",  # blank = purana password retain
        imap_sent_folder=(f.get("imap_sent_folder") or "Sent").strip(),
    )
    db.record_audit(user["username"], "MAIL_SETTINGS_UPDATED")
    return redirect(url_for("user_settings", saved="mail"))


@app.route("/settings/ai", methods=["POST"])
@login_required
def save_ai_settings():
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    user = current_user()
    f = request.form
    provider = (f.get("ai_provider") or "none").strip().lower()
    if provider not in ("none", "gemini", "anthropic"):
        provider = "none"
    db.save_ai_credentials(
        user["id"], ai_provider=provider,
        gemini_api_key=f.get("gemini_api_key") or "",  # blank = purani key retain
        anthropic_api_key=f.get("anthropic_api_key") or "",  # blank = purani key retain
    )
    db.record_audit(user["username"], "AI_SETTINGS_UPDATED")
    return redirect(url_for("user_settings", saved="ai"))


@app.route("/settings/test-mail", methods=["POST"])
@login_required
def test_mail_settings():
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    user = current_user()
    cfg = g.mail_cfg
    target = cfg.get("smtp_from_email") or cfg.get("smtp_user") or ""
    if not target:
        return redirect(url_for("user_settings", error="Pehle Mail Settings save karein, phir test bhejein."))
    try:
        msg = EmailMessage()
        msg["Subject"] = "OM Suite — Test Mail"
        msg["From"] = f"{cfg.get('smtp_from_name') or ''} <{cfg.get('smtp_from_email') or cfg.get('smtp_user')}>"
        msg["To"] = target
        msg.set_content(
            f"Ye ek test mail hai — agar ye aapko mil gayi hai to aapki Mail Settings sahi se kaam kar rahi hain.\n\n"
            f"Bheja gaya: {user['name']} ({user['username']}) ke account se."
        )
        _send_smtp(msg, [target])
    except Exception as exc:
        return redirect(url_for("user_settings", error=f"Test mail fail ho gayi: {exc}"))
    return redirect(url_for("user_settings", saved="test-mail"))


# ---------------- Diagnostics ----------------

def ai_config_status():
    provider = g.ai_cfg.get("provider", "none")
    return {
        "provider": provider,
        "enabled": provider in ("gemini", "anthropic"),
        "model": GEMINI_MODEL if provider == "gemini" else ANTHROPIC_MODEL if provider == "anthropic" else "",
    }


@app.route("/api/health", methods=["GET"])
@login_required
def api_health():
    """Ab per-request current user ki apni config dikhata hai (login required),
    kisi shared admin credential ka status nahi."""
    cfg = g.mail_cfg
    return jsonify({
        "ok": True,
        "smtp": {
            "host_configured": bool(cfg.get("smtp_host")),
            "port": cfg.get("smtp_port"),
            "user_configured": bool(cfg.get("smtp_user")),
            "from_email": cfg.get("smtp_from_email"),
        },
        "ai": ai_config_status(),
    })


# ---------------- Dashboard ----------------

@app.route("/", methods=["GET"])
@login_required
def dashboard():
    return render_template("dashboard.html", me=current_user(), ai_provider=g.ai_cfg.get("provider", "none"),
                            ai_config=ai_config_status(), mail_configured=g.mail_cfg.get("configured", False),
                            gr_auto_url=GR_AUTO_URL)


# ================== TBB PARTY MAIL — routes ==================

@app.route("/tbb", methods=["GET"])
@login_required
def tbb_index():
    has_cached_directory = os.path.exists(DIRECTORY_CACHE)
    return render_template("tbb_index.html", has_cached_directory=has_cached_directory, me=current_user())


@app.route("/tbb/upload", methods=["POST"])
@login_required
def tbb_upload():
    user_id = current_user()["id"]
    report_file = request.files.get("report_file")
    directory_file = request.files.get("directory_file")

    if not report_file or report_file.filename == "":
        return "Report/TBB sheet upload karna zaroori hai.", 400
    if not _validate_upload_extension(report_file):
        return "Sirf .xls ya .xlsx files allowed hain.", 400
    if directory_file and directory_file.filename and not _validate_upload_extension(directory_file):
        return "Directory file sirf .xls ya .xlsx honi chahiye.", 400

    udir = user_upload_dir(user_id)
    report_filename = secure_filename(report_file.filename)
    report_path = os.path.join(udir, report_filename)
    report_file.save(report_path)

    if directory_file and directory_file.filename != "":
        directory_path = os.path.join(udir, secure_filename(directory_file.filename))
        directory_file.save(directory_path)
        with open(DIRECTORY_CACHE, "wb") as f:
            with open(directory_path, "rb") as src:
                f.write(src.read())
        STATE["directory"] = {}  # cache invalidate, reload fresh below
    elif os.path.exists(DIRECTORY_CACHE):
        directory_path = DIRECTORY_CACHE
    else:
        return "Directory file bhi pehli baar upload karni hogi.", 400

    directory = parse_directory(directory_path)
    STATE["directory"] = directory
    report_branches = parse_report(report_path)
    branches = build_branches_with_email(report_branches, directory)

    for b in branches.values():
        b["ai_status"] = "pending"
        b["ai_reason"] = ""
        b["send_status"] = "pending"

    db.save_workflow_data(user_id, "tbb", branches, source_filename=report_filename)

    fingerprints = {code: compute_fingerprint("tbb", code, b.get("email") or "", tbb_build_body(b))
                     for code, b in branches.items()}
    db.sync_operations(user_id, "tbb", fingerprints)

    return redirect(url_for("tbb_preview"))


@app.route("/tbb/preview", methods=["GET"])
@login_required
def tbb_preview():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "tbb")
    operations = db.get_operations(user_id, "tbb")
    for code, b in branches.items():
        op = operations.get(code, {})
        b["op_status"] = op.get("status", "DRAFT")
        b["op_selected"] = bool(op.get("selected", 1))
        b["op_attempts"] = op.get("attempt_count", 0)
        b["op_error"] = op.get("last_error", "")
    total_gr = sum(len(b["grs"]) for b in branches.values())
    missing_email = sum(1 for b in branches.values() if not b.get("email"))
    counts = {
        "generated": len(branches),
        "queued": sum(1 for b in branches.values() if b["op_status"] == "QUEUED"),
        "skipped": sum(1 for b in branches.values() if b["op_status"] == "SKIPPED"),
        "sent": sum(1 for b in branches.values() if b["op_status"] == "SENT"),
        "failed": sum(1 for b in branches.values() if b["op_status"] == "FAILED"),
        "draft": sum(1 for b in branches.values() if b["op_status"] == "DRAFT"),
    }
    return render_template("tbb_preview.html", branches=branches, total_gr=total_gr,
                            missing_email=missing_email, ai_enabled=(g.ai_cfg.get("provider","none") != "none"),
                            me=current_user(), counts=counts)


@app.route("/api/tbb/confirm_selection", methods=["POST"])
@login_required
def api_tbb_confirm_selection():
    user_id = current_user()["id"]
    payload = request.get_json(silent=True) or {}
    selected_codes = payload.get("selected_codes", [])
    db.confirm_selection(user_id, "tbb", selected_codes)
    ops = db.get_operations(user_id, "tbb")
    return jsonify({
        "generated": len(ops),
        "selected": sum(1 for o in ops.values() if o["status"] == "QUEUED"),
        "skipped": sum(1 for o in ops.values() if o["status"] == "SKIPPED"),
    })


@app.route("/api/tbb/ai_check/<code>", methods=["POST"])
@login_required
def api_tbb_ai_check(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "tbb")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "branch not found"}), 404
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        result = {"status": "missing", "reason": "Is branch ke liye email nahi mila — manually dalein."}
    else:
        prompt = build_admin_check_prompt("TBB Party GR Verification", branch["branch_code"], branch.get("branch_name_sheet"),
                                           branch.get("region"), email, branch.get("contact_name"), branch.get("designation"),
                                           len(branch["grs"]), branch["grs"][:5])
        result = run_ai_check(prompt)
    branch["ai_status"] = result["status"]
    branch["ai_reason"] = result["reason"]
    db.save_workflow_data(user_id, "tbb", branches)
    return jsonify(result)


def _known_directory_domains():
    """STATE['directory'] mein jitni bhi emails hain unke domains ka set — typo-match ke liye reference."""
    domains = set()
    for entries in (STATE.get("directory") or {}).values():
        for e in entries:
            for em in (e.get("emails") or []):
                if "@" in em:
                    domains.add(em.split("@", 1)[1].strip().lower())
    return domains


def autofix_email_string(raw_email):
    """Whitespace/case clean-up + known-domain typo correction (e.g. 'omlogistic.co.in' -> 'omlogistics.co.in').
    Returns (fixed_email_or_None, notes[]) — None agar kuch fix karne layak nahi mila."""
    if not raw_email or not raw_email.strip():
        return None, []
    valid_domains = _known_directory_domains()
    notes = []
    changed = False
    fixed_parts = []
    for part in re.split(r"[,;\n]+", raw_email):
        p = part.strip()
        if not p:
            continue
        p2 = p.lower().replace(" ", "")
        if p2 != p:
            changed = True
        if "@" in p2:
            local, domain = p2.rsplit("@", 1)
            if valid_domains and domain not in valid_domains:
                close = difflib.get_close_matches(domain, valid_domains, n=1, cutoff=0.72)
                if close and close[0] != domain:
                    notes.append(f"'{domain}' ko '{close[0]}' se badla")
                    domain = close[0]
                    changed = True
            p2 = f"{local}@{domain}"
        fixed_parts.append(p2)
    if not fixed_parts:
        return None, []
    fixed = ", ".join(dict.fromkeys(fixed_parts))
    if not changed:
        return None, []
    return fixed, notes


@app.route("/api/tbb/ai_fix/<code>", methods=["POST"])
@login_required
def api_tbb_ai_fix(code):
    """AI Check jo 'review'/'missing'/'ai_error' de, uske liye best-effort auto-fix:
    1) Email missing ho to directory se dobara lookup karta hai.
    2) Email ho lekin galat/typo domain ho to known directory domains se match karke sudharta hai.
    3) Fix ke baad AI check dobara chalata hai aur naya status/email return karta hai.
    Kuch bhi fix na ho paye to 'unresolved' status ke saath manual review ka reason deta hai."""
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "tbb")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "branch not found"}), 404

    payload = request.get_json(silent=True) or {}
    current_email = (payload.get("email") or branch.get("email") or "").strip()
    fix_notes = []
    new_email = current_email

    if not current_email:
        directory = STATE.get("directory") or {}
        role_emails, role_names = pick_branch_role_contacts(directory, branch["branch_code"])
        if role_emails:
            new_email = ", ".join(role_emails)
            branch["contact_name"] = "; ".join(role_names)
            fix_notes.append("Email missing thi — directory se role contacts (Incharge/Billing/Delivery/DBP) dobara fetch kiye.")
        else:
            single_email, name, desig = pick_branch_email(directory, branch["branch_code"])
            if single_email:
                new_email = single_email
                branch["contact_name"] = name
                branch["designation"] = desig
                fix_notes.append("Email missing thi — directory se fallback contact fetch kiya.")
    else:
        fixed, notes = autofix_email_string(current_email)
        if fixed:
            new_email = fixed
            fix_notes.extend(notes)
            fix_notes.append("Email formatting/typo clean-up ki gayi.")

    if new_email == current_email and not fix_notes:
        return jsonify({
            "status": "unresolved",
            "reason": "Auto-fix ke liye kuch nahi mila (directory mein match ya close domain nahi mila) — email manually check karke fix karein.",
            "email": current_email,
        })

    branch["email"] = new_email
    prompt = build_admin_check_prompt(
        "TBB Party GR Verification", branch["branch_code"], branch.get("branch_name_sheet"),
        branch.get("region"), new_email, branch.get("contact_name"), branch.get("designation"),
        len(branch["grs"]), branch["grs"][:5],
    )
    result = run_ai_check(prompt)
    branch["ai_status"] = result["status"]
    combined_reason = "; ".join(fix_notes)
    if result["reason"]:
        combined_reason = f"{combined_reason} — {result['reason']}" if combined_reason else result["reason"]
    branch["ai_reason"] = combined_reason
    db.save_workflow_data(user_id, "tbb", branches)
    return jsonify({
        "status": result["status"],
        "reason": combined_reason,
        "email": new_email,
        "fixed": True,
        "notes": fix_notes,
    })


@app.route("/api/tbb/send/<code>", methods=["POST"])
@login_required
def api_tbb_send(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "tbb")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "branch not found"}), 404
    ops = db.get_operations(user_id, "tbb")
    op = ops.get(code)
    if op and op["status"] == "SENT":
        return jsonify({"status": "already_sent", "detail": "Ye mail pehle hi bheji ja chuki hai — dobara nahi bheji jayegi."})
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        return jsonify({"status": "error", "detail": "Email address missing"}), 400
    uname = current_user()["username"]
    db.mark_sending(user_id, "tbb", code)
    try:
        result_send = tbb_send_email(branch, email, payload.get("cc", ""), payload.get("bcc", ""))
        branch["send_status"] = "sent"
        db.save_workflow_data(user_id, "tbb", branches)
        db.mark_sent(user_id, "tbb", code)
        sent_copy = result_send.get("sent_copy") or {}
        if sent_copy.get("saved"):
            log_send(TBB_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "sent", username=uname)
        else:
            # Recipient ko mail chali gayi (SMTP OK), lekin "Sent" folder mein copy save nahi hui —
            # isse yehi lagta hai ki "mail bheji hi nahi", isliye ye ab clearly log + response mein flag hota hai.
            reason = sent_copy.get("reason", "IMAP not configured")
            log_send(TBB_LOG_FILE, code, branch.get("branch_name_sheet", ""), email,
                      "sent_no_sent_copy", f"Recipient ko mail deliver ho gayi lekin Sent folder copy fail: {reason}",
                      username=uname)
        return jsonify({"status": "sent", **result_send})
    except Exception as e:
        branch["send_status"] = "failed"
        db.save_workflow_data(user_id, "tbb", branches)
        db.mark_failed(user_id, "tbb", code, str(e))
        log_send(TBB_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "failed", str(e), username=uname)
        return jsonify({"status": "error", "detail": str(e)}), 500


@app.route("/api/tbb/send_all", methods=["POST"])
@login_required
def api_tbb_send_all():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "tbb")
    ops = db.get_operations(user_id, "tbb")
    payload = request.get_json(silent=True) or {}
    recipients_map = payload.get("recipients", {})  # {branch_code: {"cc": "...", "bcc": "..."}}
    results = []
    uname = current_user()["username"]
    for code, branch in branches.items():
        op = ops.get(code)
        if not op or op["status"] != "QUEUED":
            results.append({"code": code, "status": f"skipped_not_queued ({op['status'] if op else 'no-op'})"})
            continue
        email = branch.get("email")
        if not email:
            results.append({"code": code, "status": "skipped_missing_email"})
            continue
        if branch.get("ai_status") not in ("ok",):
            results.append({"code": code, "status": "skipped_needs_review", "detail": f"AI status is {branch.get('ai_status', 'pending')}. Run AI Check or use individual manual send."})
            continue
        row_recipients = recipients_map.get(code, {})
        cc = row_recipients.get("cc", "")
        bcc = row_recipients.get("bcc", "")
        db.mark_sending(user_id, "tbb", code)
        try:
            tbb_send_email(branch, email, cc, bcc)
            branch["send_status"] = "sent"
            db.mark_sent(user_id, "tbb", code)
            log_send(TBB_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "sent", username=uname)
            results.append({"code": code, "status": "sent"})
        except Exception as e:
            branch["send_status"] = "failed"
            db.mark_failed(user_id, "tbb", code, str(e))
            log_send(TBB_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "failed", str(e), username=uname)
            results.append({"code": code, "status": "failed", "detail": str(e)})
    db.save_workflow_data(user_id, "tbb", branches)
    return jsonify({"results": results})


@app.route("/api/tbb/retry_failed", methods=["POST"])
@login_required
def api_tbb_retry_failed():
    user_id = current_user()["id"]
    requeued = db.retry_failed_to_queued(user_id, "tbb")
    return jsonify({"requeued": requeued})


@app.route("/tbb/logs", methods=["GET"])
@login_required
def tbb_logs():
    if not os.path.exists(TBB_LOG_FILE):
        return "Abhi tak koi mail bheji nahi gayi."
    return send_file(TBB_LOG_FILE, as_attachment=True)


# ================== BILL GENERATED MAIL — routes ==================

@app.route("/bill", methods=["GET"])
@login_required
def bill_index():
    has_cached_directory = os.path.exists(DIRECTORY_CACHE)
    return render_template("bill_index.html", has_cached_directory=has_cached_directory, me=current_user())


@app.route("/bill/upload", methods=["POST"])
@login_required
def bill_upload():
    bill_file = request.files.get("bill_file")
    directory_file = request.files.get("directory_file")

    if not bill_file or bill_file.filename == "":
        return "Bill/report sheet upload karna zaroori hai.", 400
    if not _validate_upload_extension(bill_file):
        return "Sirf .xls ya .xlsx files allowed hain.", 400
    if directory_file and directory_file.filename and not _validate_upload_extension(directory_file):
        return "Directory file sirf .xls ya .xlsx honi chahiye.", 400

    user_id = current_user()["id"]
    udir = user_upload_dir(user_id)
    bill_filename = secure_filename(bill_file.filename)
    bill_path = os.path.join(udir, bill_filename)
    bill_file.save(bill_path)

    if directory_file and directory_file.filename != "":
        directory_path = os.path.join(udir, secure_filename(directory_file.filename))
        directory_file.save(directory_path)
        with open(DIRECTORY_CACHE, "wb") as f:
            with open(directory_path, "rb") as src:
                f.write(src.read())
        STATE["directory"] = {}
    elif os.path.exists(DIRECTORY_CACHE):
        directory_path = DIRECTORY_CACHE
    else:
        return "Directory file bhi pehli baar upload karni hogi.", 400

    directory = parse_directory(directory_path)
    STATE["directory"] = directory
    try:
        branches = bill_parser.parse_bill_report(bill_path)
    except ValueError as e:
        return str(e), 400

    from directory_parser import pick_branch_email
    for bc, b in branches.items():
        email, name, desig = pick_branch_email(directory, bc)
        b["email"] = email
        b["contact_name"] = name
        b["designation"] = desig
        b["ai_status"] = "pending"
        b["ai_reason"] = ""
        b["send_status"] = "pending"
        b["subject_preview"] = bill_parser.build_subject(b)
        b["table_html"] = bill_parser.build_table_html(b)

    db.save_workflow_data(user_id, "bill", branches, source_filename=bill_filename)

    fingerprints = {code: compute_fingerprint("bill", code, b.get("email") or "", bill_parser.build_email_text(b))
                     for code, b in branches.items()}
    db.sync_operations(user_id, "bill", fingerprints)

    return redirect(url_for("bill_preview"))


@app.route("/bill/preview", methods=["GET"])
@login_required
def bill_preview():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "bill")
    operations = db.get_operations(user_id, "bill")
    for code, b in branches.items():
        op = operations.get(code, {})
        b["op_status"] = op.get("status", "DRAFT")
        b["op_selected"] = bool(op.get("selected", 1))
        b["op_attempts"] = op.get("attempt_count", 0)
        b["op_error"] = op.get("last_error", "")
    total_bills = sum(len(b["bills"]) for b in branches.values())
    missing_email = sum(1 for b in branches.values() if not b.get("email"))
    counts = {
        "generated": len(branches),
        "queued": sum(1 for b in branches.values() if b["op_status"] == "QUEUED"),
        "skipped": sum(1 for b in branches.values() if b["op_status"] == "SKIPPED"),
        "sent": sum(1 for b in branches.values() if b["op_status"] == "SENT"),
        "failed": sum(1 for b in branches.values() if b["op_status"] == "FAILED"),
        "draft": sum(1 for b in branches.values() if b["op_status"] == "DRAFT"),
    }
    return render_template("bill_preview.html", branches=branches, total_bills=total_bills,
                            missing_email=missing_email, ai_enabled=(g.ai_cfg.get("provider","none") != "none"),
                            me=current_user(), counts=counts)


@app.route("/api/bill/confirm_selection", methods=["POST"])
@login_required
def api_bill_confirm_selection():
    user_id = current_user()["id"]
    payload = request.get_json(silent=True) or {}
    selected_codes = payload.get("selected_codes", [])
    db.confirm_selection(user_id, "bill", selected_codes)
    ops = db.get_operations(user_id, "bill")
    return jsonify({
        "generated": len(ops),
        "selected": sum(1 for o in ops.values() if o["status"] == "QUEUED"),
        "skipped": sum(1 for o in ops.values() if o["status"] == "SKIPPED"),
    })


@app.route("/api/bill/ai_check/<code>", methods=["POST"])
@login_required
def api_bill_ai_check(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "bill")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "branch not found"}), 404
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        result = {"status": "missing", "reason": "Is branch ke liye email nahi mila — manually dalein."}
    else:
        prompt = build_admin_check_prompt("Auto Bill Generated Mail", branch["branch_code"], branch.get("branch_name_sheet"),
                                           branch.get("region"), email, branch.get("contact_name"), branch.get("designation"),
                                           len(branch["bills"]), branch["bills"][:5])
        result = run_ai_check(prompt)
    branch["ai_status"] = result["status"]
    branch["ai_reason"] = result["reason"]
    db.save_workflow_data(user_id, "bill", branches)
    return jsonify(result)


@app.route("/api/bill/send/<code>", methods=["POST"])
@login_required
def api_bill_send(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "bill")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "branch not found"}), 404
    ops = db.get_operations(user_id, "bill")
    op = ops.get(code)
    if op and op["status"] == "SENT":
        return jsonify({"status": "already_sent", "detail": "Ye mail pehle hi bheji ja chuki hai — dobara nahi bheji jayegi."})
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        return jsonify({"status": "error", "detail": "Email address missing"}), 400
    email = email.strip().lower()
    if not EMAIL_RE.match(email):
        return jsonify({"status": "error", "detail": f"Invalid email address: {email}"}), 400
    uname = current_user()["username"]
    db.mark_sending(user_id, "bill", code)
    try:
        branch["email"] = email
        result_send = bill_send_email(branch, email, payload.get("cc", ""), payload.get("bcc", ""))
        branch["send_status"] = "sent"
        db.save_workflow_data(user_id, "bill", branches)
        db.mark_sent(user_id, "bill", code)
        log_send(BILL_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "sent", username=uname)
        return jsonify({"status": "sent", **result_send})
    except Exception as e:
        branch["send_status"] = "failed"
        db.save_workflow_data(user_id, "bill", branches)
        db.mark_failed(user_id, "bill", code, str(e))
        log_send(BILL_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "failed", str(e), username=uname)
        return jsonify({"status": "error", "detail": str(e)}), 500


@app.route("/api/bill/send_all", methods=["POST"])
@login_required
def api_bill_send_all():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "bill")
    ops = db.get_operations(user_id, "bill")
    payload = request.get_json(silent=True) or {}
    recipients_map = payload.get("recipients", {})
    results = []
    uname = current_user()["username"]
    for code, branch in branches.items():
        op = ops.get(code)
        if not op or op["status"] != "QUEUED":
            results.append({"code": code, "status": f"skipped_not_queued ({op['status'] if op else 'no-op'})"})
            continue
        email = branch.get("email")
        if not email:
            results.append({"code": code, "status": "skipped_missing_email"})
            continue
        if branch.get("ai_status") not in ("ok",):
            results.append({"code": code, "status": "skipped_needs_review", "detail": f"AI status is {branch.get('ai_status', 'pending')}. Run AI Check or use individual manual send."})
            continue
        row_recipients = recipients_map.get(code, {})
        cc = row_recipients.get("cc", "")
        bcc = row_recipients.get("bcc", "")
        db.mark_sending(user_id, "bill", code)
        try:
            bill_send_email(branch, email, cc, bcc)
            branch["send_status"] = "sent"
            db.mark_sent(user_id, "bill", code)
            log_send(BILL_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "sent", username=uname)
            results.append({"code": code, "status": "sent"})
        except Exception as e:
            branch["send_status"] = "failed"
            db.mark_failed(user_id, "bill", code, str(e))
            log_send(BILL_LOG_FILE, code, branch.get("branch_name_sheet", ""), email, "failed", str(e), username=uname)
            results.append({"code": code, "status": "failed", "detail": str(e)})
    db.save_workflow_data(user_id, "bill", branches)
    return jsonify({"results": results})


@app.route("/api/bill/retry_failed", methods=["POST"])
@login_required
def api_bill_retry_failed():
    user_id = current_user()["id"]
    requeued = db.retry_failed_to_queued(user_id, "bill")
    return jsonify({"requeued": requeued})


@app.route("/bill/logs", methods=["GET"])
@login_required
def bill_logs():
    if not os.path.exists(BILL_LOG_FILE):
        return "Abhi tak koi mail bheji nahi gayi."
    return send_file(BILL_LOG_FILE, as_attachment=True)


@app.route("/mail/history", methods=["GET"])
@login_required
def mail_history():
    rows = []
    for module, path in (("TBB", TBB_LOG_FILE), ("BILL", BILL_LOG_FILE)):
        if os.path.exists(path):
            with open(path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    row["module"] = module; rows.append(row)
    rows.sort(key=lambda r: r.get("timestamp", ""), reverse=True)
    return render_template("mail_history.html", rows=rows, me=current_user())


# ==================== THUNDERBIRD BRIDGE ====================
# Extension apni HTTP requests mein X-Thunderbird-Token aur X-Thunderbird-User headers
# bhejta hai — normal login session cookie use nahi hota, isliye ye endpoints alag
# se authenticate hote hain (session-based @login_required se nahi).

def bridge_auth():
    """Request headers se user nikalta hai. None return karta hai agar auth fail ho."""
    token = request.headers.get("X-Thunderbird-Token", "")
    username = request.headers.get("X-Thunderbird-User", "")
    return db.get_user_by_bridge_auth(username, token)


@app.route("/api/bridge/status", methods=["GET"])
def api_bridge_status():
    user = bridge_auth()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    pending = len([j for j in db.list_recent_bridge_jobs(user["id"], limit=50) if j["status"] in ("pending", "delivered")])
    return jsonify({"ok": True, "username": user["username"], "pending_jobs": pending})


@app.route("/api/bridge/poll", methods=["GET"])
def api_bridge_poll():
    user = bridge_auth()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    limit = int(request.args.get("limit", 5))
    jobs = db.poll_bridge_jobs(user["id"], limit=limit)
    return jsonify({"jobs": jobs})


@app.route("/api/bridge/job/<int:job_id>/result", methods=["POST"])
def api_bridge_job_result(job_id):
    user = bridge_auth()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    if payload.get("status") == "error":
        db.store_bridge_job_result(job_id, user["id"], result=None, status="error", error=payload.get("error", ""))
    else:
        db.store_bridge_job_result(job_id, user["id"], result=payload.get("result"), status="done")
    return jsonify({"ok": True})


@app.route("/api/bridge/attachment/<module>/<code>", methods=["GET"])
def api_bridge_attachment(module, code):
    """Extension send-job ke waqt attachment fetch karta hai (agar payload mein url diya ho)."""
    user = bridge_auth()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if module not in ("tbb", "bill"):
        return jsonify({"error": "invalid module"}), 400
    branches = db.get_workflow_data(user["id"], module)
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "not found"}), 404
    if module == "tbb":
        csv_bytes = tbb_build_attachment_csv(branch)
        filename = f"TBB_{code}.csv"
    else:
        csv_bytes = bill_parser.build_attachment_csv(branch)
        filename = f"Bill_{code}.csv"
    return (csv_bytes, 200, {"Content-Type": "text/csv", "Content-Disposition": f"attachment; filename={filename}"})


@app.route("/bridge", methods=["GET"])
@login_required
def bridge_settings():
    user = current_user()
    token = db.get_or_create_bridge_token(user["id"])
    jobs = db.list_recent_bridge_jobs(user["id"], limit=20)
    return render_template("bridge.html", me=user, token=token, jobs=jobs,
                            csrf_token=get_csrf_token(), api_base=request.host_url.rstrip("/"))


@app.route("/bridge/regenerate", methods=["POST"])
@login_required
def bridge_regenerate():
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    db.regenerate_bridge_token(current_user()["id"])
    return redirect(url_for("bridge_settings"))


@app.route("/api/bridge/job/<int:job_id>/status", methods=["GET"])
@login_required
def api_bridge_job_status(job_id):
    """Web UI (session-authenticated) is se job ka result poll karta hai — jaise
    'match' job complete hua ki nahi, taaki recipients auto-fill ho sakein."""
    job = db.get_bridge_job(job_id, current_user()["id"])
    if not job:
        return jsonify({"error": "not found"}), 404
    return jsonify(job)


def _extract_field_values(rows, keyword):
    """Rows (list of dicts) mein se un columns ki values nikalta hai jinke naam mein
    'keyword' aata hai (jaise 'party' -> 'Party Name' column). Bridge match job ke
    liye party/bill numbers dhoondne ke kaam aata hai."""
    values = []
    seen = set()
    for row in rows:
        for k, v in row.items():
            if keyword.lower() in str(k).lower() and v not in (None, ""):
                sv = str(v).strip()
                if sv and sv not in seen:
                    seen.add(sv)
                    values.append(sv)
    return values


@app.route("/previous-mail", methods=["GET"])
@login_required
def previous_mail_index():
    return render_template("previous_index.html", me=current_user())


@app.route("/previous-mail/analyze", methods=["POST"])
@login_required
def previous_mail_analyze():
    user_id = current_user()["id"]
    udir = user_upload_dir(user_id)

    eml_file = request.files.get("eml_file")
    data_file = request.files.get("data_file")
    if not data_file or data_file.filename == "":
        return "Naya data sheet upload karna zaroori hai.", 400
    if not _validate_upload_extension(data_file) and not data_file.filename.lower().endswith(".csv"):
        return "Data sheet .xls, .xlsx ya .csv honi chahiye.", 400

    data_path = os.path.join(udir, secure_filename(data_file.filename))
    data_file.save(data_path)

    if eml_file and eml_file.filename:
        old_mail = previous_mail.parse_eml_bytes(eml_file.read())
    else:
        old_mail = previous_mail.parse_manual_fields(
            request.form.get("old_to", ""), request.form.get("old_cc", ""),
            request.form.get("old_bcc", ""), request.form.get("old_subject", ""),
            request.form.get("old_body", ""),
        )
        if not old_mail["subject"] and not old_mail["body"]:
            return "Purani mail ka .eml upload karein YA subject/body manually likhein.", 400

    df = previous_mail.parse_new_data(data_path)
    columns = list(df.columns)
    if not columns:
        return "Data sheet mein koi column nahi mila.", 400
    sample_row = {c: ("" if pd.isna(df.iloc[0][c]) else df.iloc[0][c]) for c in columns}

    ai_result = run_previous_template_ai(old_mail["subject"], old_mail["body"], sample_row, columns)

    default_key_col = next((c for c in columns if "branch" in c.lower()), columns[0])

    draft = {
        "old_mail": old_mail,
        "data_path": data_path,
        "columns": columns,
        "sample_row": sample_row,
        "subject_template": ai_result["subject_template"],
        "body_template": ai_result["body_template"],
        "ai_confidence": ai_result["confidence"],
        "ai_notes": ai_result["notes"],
        "key_col": default_key_col,
    }
    db.save_workflow_data(user_id, "previous_draft", draft)
    return redirect(url_for("previous_mail_review"))


@app.route("/previous-mail/review", methods=["GET"])
@login_required
def previous_mail_review():
    user_id = current_user()["id"]
    draft = db.get_workflow_data(user_id, "previous_draft")
    if not draft:
        return redirect(url_for("previous_mail_index"))
    return render_template("previous_review.html", me=current_user(), draft=draft, csrf_token=get_csrf_token())


@app.route("/previous-mail/confirm", methods=["POST"])
@login_required
def previous_mail_confirm():
    user_id = current_user()["id"]
    if not validate_csrf(request.form.get("csrf_token")):
        abort(400)
    draft = db.get_workflow_data(user_id, "previous_draft")
    if not draft:
        return redirect(url_for("previous_mail_index"))

    subject_template = request.form.get("subject_template", draft["subject_template"])
    body_template = request.form.get("body_template", draft["body_template"])
    key_col = request.form.get("key_col", draft["key_col"])
    default_to = request.form.get("to", draft["old_mail"].get("to", ""))
    default_cc = request.form.get("cc", draft["old_mail"].get("cc", ""))
    default_bcc = request.form.get("bcc", draft["old_mail"].get("bcc", ""))

    df = previous_mail.parse_new_data(draft["data_path"])
    if key_col not in df.columns:
        return f"Column '{key_col}' data sheet mein nahi mila.", 400
    groups = previous_mail.group_by_key(df, key_col)

    branches = {}
    for key, rows in groups.items():
        subject = previous_mail.render_template(subject_template, rows[0])
        body = previous_mail.render_template(body_template, rows[0])
        if len(rows) > 1:
            body += "\n\n---\nIs group ke saare rows:\n" + previous_mail.render_rows_table_text(rows, draft["columns"])
        branches[key] = {
            "branch_code": key,
            "key": key,
            "columns": draft["columns"],
            "rows": rows,
            "subject": subject,
            "body": body,
            "email": default_to,
            "contact_name": "",
            "ai_status": "pending",
            "ai_reason": "",
            "send_status": "pending",
        }

    db.save_workflow_data(user_id, "previous", branches, source_filename=os.path.basename(draft["data_path"]))
    fingerprints = {key: compute_fingerprint("previous", key, b.get("email") or "", b["subject"] + b["body"])
                     for key, b in branches.items()}
    db.sync_operations(user_id, "previous", fingerprints)

    return redirect(url_for("previous_mail_preview"))


@app.route("/previous-mail/preview", methods=["GET"])
@login_required
def previous_mail_preview():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "previous")
    operations = db.get_operations(user_id, "previous")
    for code, b in branches.items():
        op = operations.get(code, {})
        b["op_status"] = op.get("status", "DRAFT")
        b["op_selected"] = bool(op.get("selected", 1))
        b["op_attempts"] = op.get("attempt_count", 0)
        b["op_error"] = op.get("last_error", "")
    counts = {
        "generated": len(branches),
        "queued": sum(1 for b in branches.values() if b["op_status"] == "QUEUED"),
        "skipped": sum(1 for b in branches.values() if b["op_status"] == "SKIPPED"),
        "sent": sum(1 for b in branches.values() if b["op_status"] == "SENT"),
        "failed": sum(1 for b in branches.values() if b["op_status"] == "FAILED"),
        "draft": sum(1 for b in branches.values() if b["op_status"] == "DRAFT"),
    }
    return render_template("previous_preview.html", branches=branches, counts=counts,
                            ai_enabled=(g.ai_cfg.get("provider","none") != "none"), me=current_user())


@app.route("/api/previous/confirm_selection", methods=["POST"])
@login_required
def api_previous_confirm_selection():
    user_id = current_user()["id"]
    payload = request.get_json(silent=True) or {}
    db.confirm_selection(user_id, "previous", payload.get("selected_codes", []))
    ops = db.get_operations(user_id, "previous")
    return jsonify({"generated": len(ops), "selected": sum(1 for o in ops.values() if o["status"] == "QUEUED"),
                     "skipped": sum(1 for o in ops.values() if o["status"] == "SKIPPED")})


@app.route("/api/previous/ai_check/<code>", methods=["POST"])
@login_required
def api_previous_ai_check(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "previous")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "not found"}), 404
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        result = {"status": "missing", "reason": "Email address nahi hai — manually dalein."}
    else:
        prompt = build_admin_check_prompt("Previous Mail & Format", branch["key"], branch["key"], "", email,
                                           "", "", len(branch["rows"]), branch["rows"][:5])
        result = run_ai_check(prompt)
        if result["status"] == "skipped":
            result["status"] = "ok"
    branch["ai_status"] = result["status"]
    branch["ai_reason"] = result["reason"]
    db.save_workflow_data(user_id, "previous", branches)
    return jsonify(result)


@app.route("/api/previous/send/<code>", methods=["POST"])
@login_required
def api_previous_send(code):
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "previous")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "not found"}), 404
    ops = db.get_operations(user_id, "previous")
    op = ops.get(code)
    if op and op["status"] == "SENT":
        return jsonify({"status": "already_sent", "detail": "Ye mail pehle hi bheji ja chuki hai."})
    payload = request.get_json(silent=True) or {}
    email = payload.get("email") or branch.get("email") or ""
    if not email:
        return jsonify({"status": "error", "detail": "Email address missing"}), 400
    uname = current_user()["username"]
    db.mark_sending(user_id, "previous", code)
    try:
        previous_send_email(branch, email, payload.get("cc", ""), payload.get("bcc", ""))
        branch["send_status"] = "sent"
        db.save_workflow_data(user_id, "previous", branches)
        db.mark_sent(user_id, "previous", code)
        log_send(PREVIOUS_LOG_FILE, code, code, email, "sent", username=uname)
        return jsonify({"status": "sent"})
    except Exception as e:
        branch["send_status"] = "failed"
        db.save_workflow_data(user_id, "previous", branches)
        db.mark_failed(user_id, "previous", code, str(e))
        log_send(PREVIOUS_LOG_FILE, code, code, email, "failed", str(e), username=uname)
        return jsonify({"status": "error", "detail": str(e)}), 500


@app.route("/api/previous/send_all", methods=["POST"])
@login_required
def api_previous_send_all():
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "previous")
    ops = db.get_operations(user_id, "previous")
    payload = request.get_json(silent=True) or {}
    recipients_map = payload.get("recipients", {})
    results = []
    uname = current_user()["username"]
    for code, branch in branches.items():
        op = ops.get(code)
        if not op or op["status"] != "QUEUED":
            results.append({"code": code, "status": f"skipped_not_queued ({op['status'] if op else 'no-op'})"})
            continue
        email = branch.get("email")
        if not email:
            results.append({"code": code, "status": "skipped_missing_email"})
            continue
        if branch.get("ai_status") not in ("ok",):
            results.append({"code": code, "status": "skipped_needs_review"})
            continue
        row_recipients = recipients_map.get(code, {})
        db.mark_sending(user_id, "previous", code)
        try:
            previous_send_email(branch, email, row_recipients.get("cc", ""), row_recipients.get("bcc", ""))
            branch["send_status"] = "sent"
            db.mark_sent(user_id, "previous", code)
            log_send(PREVIOUS_LOG_FILE, code, code, email, "sent", username=uname)
            results.append({"code": code, "status": "sent"})
        except Exception as e:
            branch["send_status"] = "failed"
            db.mark_failed(user_id, "previous", code, str(e))
            log_send(PREVIOUS_LOG_FILE, code, code, email, "failed", str(e), username=uname)
            results.append({"code": code, "status": "failed", "detail": str(e)})
    db.save_workflow_data(user_id, "previous", branches)
    return jsonify({"results": results})


@app.route("/api/previous/retry_failed", methods=["POST"])
@login_required
def api_previous_retry_failed():
    user_id = current_user()["id"]
    requeued = db.retry_failed_to_queued(user_id, "previous")
    return jsonify({"requeued": requeued})


@app.route("/api/previous/find_via_thunderbird/<code>", methods=["POST"])
@login_required
def api_previous_find_via_thunderbird(code):
    """Bridge 'match' job enqueue karta hai — Thunderbird extension isse poll karke
    Sent folder mein is branch/party/bill se milti-julti purani mail dhoondega."""
    user_id = current_user()["id"]
    branches = db.get_workflow_data(user_id, "previous")
    branch = branches.get(code)
    if not branch:
        return jsonify({"error": "not found"}), 404
    party_names = _extract_field_values(branch["rows"], "party")
    bill_nos = _extract_field_values(branch["rows"], "bill")
    payload = {
        "branch_code": branch["key"], "branch_name": branch["key"],
        "party_names": party_names, "bill_nos": bill_nos,
        "subject": branch.get("subject", ""),
    }
    job_id = db.enqueue_bridge_job(user_id, "match", payload)
    return jsonify({"job_id": job_id})


if __name__ == "__main__":
    app.run(debug=True, port=5000, use_reloader=False)
