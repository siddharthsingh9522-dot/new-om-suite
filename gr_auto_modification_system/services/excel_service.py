"""
Excel handling service: read uploaded workbooks, detect the GR column,
and generate the final downloadable multi-sheet report.
"""
import logging
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import settings
from utils.validators import guess_gr_column, is_valid_gr_number

logger = logging.getLogger("gr_auto_mod.excel_service")


class ExcelParseError(Exception):
    pass


def list_sheet_names(filepath: str):
    try:
        xl = pd.ExcelFile(filepath)
        return xl.sheet_names
    except Exception as exc:
        raise ExcelParseError(f"Could not read Excel file: {exc}") from exc


def load_sheet(filepath: str, sheet_name: str = None):
    """Load a single sheet as a DataFrame (all columns as string/object)."""
    try:
        if sheet_name:
            df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
        else:
            df = pd.read_excel(filepath, dtype=str)
    except Exception as exc:
        raise ExcelParseError(f"Could not parse sheet: {exc}") from exc

    df.columns = [str(c).strip() for c in df.columns]
    return df


def analyze_gr_column(df: pd.DataFrame, chosen_column: str = None):
    """
    Detect (or use the user-chosen) GR column and compute summary stats:
    total rows, valid GR count, empty rows, duplicates, invalid values.
    """
    column = chosen_column or guess_gr_column(list(df.columns))
    if not column or column not in df.columns:
        return {
            "column": None,
            "confident": False,
            "total_rows": len(df),
            "valid_count": 0,
            "empty_count": 0,
            "duplicate_count": 0,
            "invalid_count": 0,
            "gr_values": [],
        }

    raw_values = df[column].tolist()
    cleaned = [str(v).strip() if v is not None and str(v).strip().lower() != "nan" else "" for v in raw_values]

    empty_count = sum(1 for v in cleaned if v == "")
    valid_values = [v for v in cleaned if v and is_valid_gr_number(v)]
    invalid_count = sum(1 for v in cleaned if v and not is_valid_gr_number(v))

    seen = set()
    duplicates = set()
    for v in valid_values:
        if v in seen:
            duplicates.add(v)
        seen.add(v)

    return {
        "column": column,
        "confident": chosen_column is not None or guess_gr_column(list(df.columns)) == column,
        "total_rows": len(df),
        "valid_count": len(valid_values),
        "empty_count": empty_count,
        "duplicate_count": len(duplicates),
        "invalid_count": invalid_count,
        "gr_values": cleaned,
    }


def generate_report(batch, items, output_dir: str = None) -> str:
    """
    Generate the final downloadable Excel report with SUMMARY / SUCCESS /
    FAILED / SKIPPED / ALREADY_APPLIED sheets. Returns the output filepath.
    """
    output_dir = output_dir or settings.EXPORT_FOLDER
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    columns = [
        "Batch ID", "GR/CN Number", "Original Party Code", "New Party Code",
        "Party Name", "Original Remark", "Requested New Remark", "Final Remark",
        "Status", "Error Message", "Attempt Count", "Processed At",
    ]

    def write_sheet(ws, rows):
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        for col_idx, _ in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    def row_for(item):
        return [
            batch.batch_id,
            item.gr_number,
            item.existing_party_code,
            item.new_party_code,
            item.party_name,
            item.existing_remark,
            item.new_remark,
            item.final_remark,
            item.status,
            item.last_error,
            item.attempts,
            item.updated_at.isoformat() if item.updated_at else "",
        ]

    ws_summary = wb.active
    ws_summary.title = "SUMMARY"

    # Always recompute counts live from the actual items rather than
    # trusting batch.success_count/failed_count/etc - those counters are
    # only written once execute_batch() finishes, so a report downloaded
    # WHILE a batch is still RUNNING would otherwise show stale zeros even
    # though individual rows already have real statuses.
    live_success = sum(1 for i in items if i.status == "SUCCESS")
    live_failed = sum(1 for i in items if i.status in ("FAILED", "ERROR", "INVALID_CN", "INVALID_PARTY"))
    live_skipped = sum(1 for i in items if i.status == "SKIPPED")
    live_already_applied = sum(1 for i in items if i.status == "ALREADY_APPLIED")

    summary_rows = [
        ["Batch ID", batch.batch_id],
        ["Source Filename", batch.source_filename or ""],
        ["Total GR", batch.total_gr],
        ["Selected GR", batch.selected_gr],
        ["Success", live_success],
        ["Failed", live_failed],
        ["Skipped", live_skipped],
        ["Already Applied", live_already_applied],
        ["Status", batch.status],
        ["Dry Run", batch.dry_run],
        ["Common Remark", batch.common_remark or ""],
        ["Detected Party Code", batch.detected_party_code or ""],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    buckets = {
        "SUCCESS": [i for i in items if i.status == "SUCCESS"],
        "FAILED": [i for i in items if i.status in ("FAILED", "ERROR", "INVALID_CN", "INVALID_PARTY")],
        "SKIPPED": [i for i in items if i.status == "SKIPPED"],
        "ALREADY_APPLIED": [i for i in items if i.status == "ALREADY_APPLIED"],
    }

    for sheet_name, rows in buckets.items():
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, [row_for(i) for i in rows])

    filename = f"{batch.batch_id}_report.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def generate_template(output_dir: str = None) -> str:
    """Generate a sample Excel template that users can fill in for bulk upload."""
    output_dir = output_dir or settings.EXPORT_FOLDER
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "GR_LIST"
    ws.append(["GR NO"])
    for sample in ["1603261001552", "1603261001553", "1603261001554"]:
        ws.append([sample])
    ws.column_dimensions["A"].width = 25

    filepath = os.path.join(output_dir, "gr_bulk_upload_template.xlsx")
    wb.save(filepath)
    return filepath


def generate_module_report(module: dict, batch, items, output_dir: str = None) -> str:
    """
    Generate the downloadable Excel report for one of the generic-engine
    modules (Consignor / Consignee / Freight Mode / Transport Mode).
    Mirrors generate_report() but uses ModuleBatchItem's generic field
    names and includes the pre-execution columns (auto-generated vs
    user-edited final remark, change type, selected) requested for the
    Review & Confirm center's "Download Pre-Modification Report".
    """
    output_dir = output_dir or settings.EXPORT_FOLDER
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    value_label = module.get("value_label", "Value")
    columns = [
        "Batch ID", "GR/CN Number", f"Current {value_label}", f"New {value_label}",
        f"{value_label} Name/Label", "Existing Remark", "Common New Remark",
        "Auto Generated Final Remark", "User Final Remark", "Final Remark Used",
        "Change Type", "Status", "Selected", "Error Message", "Attempt Count", "Processed At",
    ]

    def write_sheet(ws, rows):
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        for col_idx, _ in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    def row_for(item):
        return [
            batch.batch_id, item.gr_number, item.existing_value, item.new_value,
            item.new_value_label or item.existing_value_label, item.existing_remark,
            item.new_remark, item.auto_final_remark, item.user_final_remark, item.final_remark,
            item.change_type, item.status, item.selected, item.last_error, item.attempts,
            item.updated_at.isoformat() if item.updated_at else "",
        ]

    ws_summary = wb.active
    ws_summary.title = "SUMMARY"
    live_success = sum(1 for i in items if i.status in ("SUCCESS", "VERIFIED_SUCCESS"))
    live_failed = sum(1 for i in items if i.status in ("FAILED", "ERROR", "INVALID_CN", "INVALID_VALUE", "VERIFICATION_FAILED"))
    live_skipped = sum(1 for i in items if i.status == "SKIPPED")
    live_already_applied = sum(1 for i in items if i.status == "ALREADY_APPLIED")
    live_selected = sum(1 for i in items if i.selected)

    summary_rows = [
        ["Batch ID", batch.batch_id],
        ["Module", module.get("display_name")],
        ["Source Filename", batch.source_filename or ""],
        ["Total GR", batch.total_gr],
        ["Selected GR", live_selected],
        ["Success", live_success],
        ["Failed", live_failed],
        ["Skipped", live_skipped],
        ["Already Applied", live_already_applied],
        ["Status", batch.status],
        ["Dry Run", batch.dry_run],
        ["Common Remark", batch.common_remark or ""],
        [f"Common New {value_label}", batch.common_new_value or ""],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 50

    buckets = {
        "ALL_RECORDS": items,
        "SUCCESS": [i for i in items if i.status in ("SUCCESS", "VERIFIED_SUCCESS")],
        "FAILED": [i for i in items if i.status in ("FAILED", "ERROR", "INVALID_CN", "INVALID_VALUE", "VERIFICATION_FAILED")],
        "SKIPPED": [i for i in items if i.status == "SKIPPED"],
        "ALREADY_APPLIED": [i for i in items if i.status == "ALREADY_APPLIED"],
    }
    for sheet_name, rows in buckets.items():
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, [row_for(i) for i in rows])

    filename = f"{batch.batch_id}_{module.get('key', 'module')}_report.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
